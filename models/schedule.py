from enum import Enum
from io import BytesIO
from typing import  List # For type hinting
from datetime import datetime
from lib.string_safety import secureString
from models.lecturer import Lecturer
from models.lesson import DayOfWeek, Lesson, strToType
from models.group import Group
from models.location import Location

def dayToDayNum(day: str):
    if  "Poniedziałek" in day:
        return 1
    elif "Wtorek" in day:
        return 2
    elif "Środa" in day:
        return 3
    elif "Czwartek" in day:
        return 4
    elif "Piątek" in day:
        return 5
    elif "Sobota" in day:
        return 6
    elif "Niedziela" in day:
        return 7
    else:
        raise Exception("NO_DAY_FOUND")

class Degree(Enum):
    ENGINEER = "inż"
    MASTER = "mgr"
    BACHELOR = "lic"

def strToDegree(string:str):
    if "inż" in string.lower():
        return Degree.ENGINEER
    elif "mgr" in string.lower() or "mag" in string.lower():
        return Degree.MASTER
    elif "lic" in string.lower():
        return Degree.BACHELOR
    else:
        raise Exception("NO_DEGREE_FOUND")

class Mode(Enum):
    STACIONARY = "ST"
    NONSTACIONARY = "ZO"

def strToMode(string:str):
    if "st" in string.lower():
        return Mode.STACIONARY
    elif "z" in string.lower() and "o" in string.lower():
        return Mode.NONSTACIONARY
    else:
        raise Exception("NO_MODE_FOUND")

class Schedule:
    def __init__(self, faculty:str="", field:str="", degree:Degree=Degree.ENGINEER, mode:Mode = Mode.STACIONARY, year:int=0, semester:int=0, lessons:List[Lesson]=[]):
        self.faculty = faculty
        self.field = field
        self.degree:Degree = degree
        self.mode:Mode = mode
        self.year = year
        self.semester = semester
        self.groups:List[Group] = []
        self.lessons:List[Lesson] = lessons

    def name(self):
        return self.field + " R" + str(self.year) + "S" + str(self.semester) + " " + self.mode.name + " " + self.degree.name

    @classmethod
    def get_timetables_from_xlsx_data_openpyxl(cls, data: BytesIO):
        from openpyxl import load_workbook, cell
        from openpyxl.worksheet.worksheet import Worksheet
        from openpyxl.cell.cell import Cell

        def parse_sheet(sheet: Worksheet):
            schedule:Schedule = Schedule(lessons=[])
            schedule_data = str(sheet.cell(row=1, column=3).value).split(sep=",")

            # Get schedule data from first rows
            schedule.field = schedule_data[0].strip()
            schedule.degree = strToDegree(schedule_data[1].strip())
            schedule.year = schedule_data[2].split("Rok")[1].strip()
            schedule.semester = schedule_data[3].split("Semestr")[1].strip()
            schedule.mode = Mode.NONSTACIONARY
            schedule.faculty = "WZIM"

            time_row:tuple[Cell,...] = ()
            i:int = 2
            # Iterate over next rows
            while str(sheet[i+1][1].value) != 'None':
                i += 1
                # Check if this is time row
                if sheet[i][1].value == "Grupy":
                    time_row = sheet[i]
                    continue

                smallSchedule = parse_row(sheet, i, time_row)
                schedule.lessons.extend(smallSchedule.lessons)
                existingGroupNames = [group.name for group in schedule.groups]
                for group in smallSchedule.groups:
                    if group.name not in existingGroupNames:
                        schedule.groups.append(group)
            return schedule

        def parse_row(worksheet:Worksheet, n:int, time_row:tuple[Cell,...]):
            schedule = Schedule(lessons=[])
            row = worksheet[n]
            max_col = len(row)
            group = Group(str(row[1].value).strip())
            schedule.groups.append(group)

            # Find day info
            temp = n
            while type(worksheet[temp][0]) == cell.MergedCell:
                temp -= 1

            dayWithInfo = str(worksheet[temp][0].value).strip()

            if "Piątek" in dayWithInfo:
                day = DayOfWeek.FRIDAY
            elif "Sobota" in dayWithInfo:
                day = DayOfWeek.SATURDAY
            elif "Niedziela" in dayWithInfo:
                day = DayOfWeek.SUNDAY
            else:
                raise Exception("NO_DAY_FOUND")

            j = 1
            while j+1 < max_col:
                j += 1
                lesson = Lesson(day=day, lecturers=[], groups=[group])
                temp = n

                # Get to the first non-merged cell
                while type(worksheet[temp][j]) == cell.MergedCell:
                    temp -= 1

                if str(worksheet[temp][j].value) == "None":
                    continue

                temprow = worksheet[temp]
                lesson.startTime = time_row[j].value

                # Save lesson data for later
                lessonRaw = str(temprow[j].value).strip()

                # Go back to the propper row and the rest of the lesson
                while (j+1 < max_col) and (type(worksheet[n][j+1]) == cell.MergedCell):
                    j += 1
                lesson.endTime = time_row[j].value

                # Parse lesson data
                lessonRaw = lessonRaw.split(",")
                lesson.name = lessonRaw[0].split("(")[0].strip()

                for text in lessonRaw:
                    if text == "WARUNEK":
                        continue
                    elif "(" in text:
                        lesson.type = strToType(string=text.split("(")[1].split(")")[0].strip())
                    elif "[" in text:
                        text = text.split("[")[1].strip()
                        if lesson.location == None:
                            lesson.location = Location()
                        if "]" in text:
                            lesson.location.building = text.split("b")[1][1:].split("]")[0].strip()
                            text = text.split("b")[0].strip()
                        if "/" in text:
                            lesson.location.floor = int(text.split("s")[1][1:].split("/")[0].strip())
                            lesson.location.classroom = text.split("s")[1][1:].split("/")[1].strip()
                        else:
                            if "s" not in text.lower():
                                lesson.location.classroom = text.strip().removesuffix(".")
                            else:
                                lesson.location.classroom = text.split("s")[1][1:].strip()

                    elif "b." in text:
                        if lesson.location == None:
                            lesson.location = Location()
                        lesson.location.building = text.split("b.")[1].split("]")[0].strip()
                    else:
                        isTeacher = True
                        for word in text.strip().split(" "):
                            if not word[0].isupper() or word[1].isupper() or any(char.isdigit() for char in word):
                                isTeacher = False
                                break
                        if isTeacher:
                            lecturerData = text.strip().split(" ")
                            lecturer = Lecturer()
                            # Name is everything except the last word
                            lecturer.name = " ".join(lecturerData[:-1])
                            lecturer.surname = lecturerData[-1].strip()
                            lesson.lecturers.append(lecturer)

                if not "]" in lessonRaw[-1] and len(lessonRaw) > 1:
                    lesson.comment = lessonRaw[-1].strip().strip(".")

                schedule.lessons.append(lesson)
            return schedule

        plans:List[Schedule] = []
        workbook = load_workbook(data, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            parsed_sheet = parse_sheet(sheet)
            plans.append(parsed_sheet)

        return plans



    @classmethod
    def get_timetable_from_txt_data(cls, data: List[str]):
        # 10.10.2023 22:34
        # WZIM, 2023, Jesień , ST, Inf, inż, R4, S7, gr1, ISI-1 ;
        lessons:List[Lesson] = []
        schedule = Schedule()
        # Get name
        scheduleInfo = data[1].strip().split(",")
        schedule.faculty = scheduleInfo[0].strip() # WZIM
        schedule.year = scheduleInfo[1].strip() # 2023
        schedule.mode = strToMode(scheduleInfo[3].strip()) # ST
        schedule.field = scheduleInfo[4].strip() # Inf
        if schedule.field == "Inf":
            schedule.field = "Informatyka"
        schedule.degree = strToDegree(scheduleInfo[5].strip()) # inż
        schedule.year = int(scheduleInfo[6].strip().removeprefix("R").strip()) # R4
        schedule.semester = int(scheduleInfo[7].strip().removeprefix("S").strip()) # S7
        group = Group(scheduleInfo[8].strip().removeprefix("gr").strip()) # gr1
        schedule.groups.append(group)

        split_filter = "------------------------------------------------"
        blocks: List[List[str]] = []
        # Divide file into blocks of lessons
        block:List[str] = []
        for line in data:
            if line == split_filter:
                blocks.append(block)
                block = []
            else:
                block.append(line)
        blocks.append(block)

        # Iterate over blocks
        for block in blocks:
            lesson = Lesson(groups=[group])
            if "ZJ" not in block[0]:
                continue
            elif "end." in block:
                break

            # Parse text
            lesson.num = int(block[0].strip("ZJ_").strip())
            lesson.name = block[1].split("[")[0].strip()
            lesson.type = strToType(block[1].split("[")[1].split("]")[0].strip())
            # lesson.day = int(block[2].split(",")[0].strip().removeprefix("d").strip())
            lesson.day = DayOfWeek.fromInt(int(block[2].split(",")[0].strip().removeprefix("d").strip()))
            timeStartText = block[2].split(",")[1].split("-")[0].strip()
            lesson.startTime = datetime.strptime(timeStartText, "%H:%M").time()
            timeEndText = block[2].split(",")[1].split("-")[1].strip()
            lesson.endTime = datetime.strptime(timeEndText, "%H:%M").time()

            if not "zdaln" in block[3]:
                lesson.location = Location()
                # floor/classroom, building
                locationText = block[3].split(",")
                if len(locationText) > 1:
                    lesson.location.building = locationText[1].split("/")[0].strip()
                else:
                    lesson.location.building = "34"

                if "/" in locationText[0]:
                    lesson.location.floor = int(locationText[0].split("/")[0].strip())
                    lesson.location.classroom = locationText[0].split("/")[1].strip()
                else:
                    lesson.location.classroom = locationText[0].strip()

            lecturer = Lecturer()
            lecturerText = block[4].split(" ")
            lecturer.name = " ".join(lecturerText[:-1])
            lecturer.surname = lecturerText[-1].strip()
            lesson.lecturers = [lecturer]
            lesson.comment = block[5].removeprefix("U:").strip()

            # Add lesson to lessons list
            lessons.append(lesson)
        schedule.lessons = lessons
        return schedule

    @classmethod
    def get_timetables_from_file(cls, filename: str) -> List["Schedule"]:
        if filename.endswith(".txt"):
            with open(filename, "r") as f:
                return [cls.get_timetable_from_txt_data(f.read().splitlines())]
        elif filename.endswith(".xlsx"):
            with open(filename, "rb") as f:
                return cls.get_timetables_from_xlsx_data_openpyxl(BytesIO(f.read()))
        else:
            return []


    def to_map(self):
        return {
            "name": secureString(self.name())
        }

    def __str__(self):
        return f"Faculty: {self.faculty}\nField: {self.field}\nDegree: {self.degree}\nMode: {self.mode}\nYear: {self.year}\nSemester: {self.semester}\nLessons: {self.lessons}"
