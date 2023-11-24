from io import BytesIO
import io
from types import CellType
from typing import List # For type hinting

from io import BytesIO
from typing import List

class File:
    def __init__(self, name='', content=b''):
        self.name = name
        self.content = content

    def save(self, folder, name=None):
        import os
        if not os.path.isdir(folder):
            os.mkdir(folder)

        with open(os.path.join(folder, name or self.name), 'wb') as f:
            f.write(self.content)

    def __str__(self):
        return f"Name: {self.name}\nContent: {self.content}"

class EmailMessage:
    def __init__(self, subject='', sender='', recipient='', date='', body='', attachments=None):
        if attachments is None:
            attachments = []
        self.subject = subject
        self.sender = sender
        self.recipient = recipient
        self.date = date
        self.body = body
        self.attachments: List[File] = attachments

    @classmethod
    def from_raw_email(cls, raw_email):
        def extract_body(parsed_email):
            body = ''
            if parsed_email.is_multipart():
                for part in parsed_email.walk():
                    ctype = part.get_content_type()
                    cdispo = str(part.get('Content-Disposition'))

                    # skip any text/plain (txt) attachments
                    if ctype == 'text/plain' and 'attachment' not in cdispo:
                        body = part.get_payload(decode=True)
                        break
            # not multipart - i.e. plain text, no attachments, keeping fingers crossed
            else:
                body = parsed_email.get_payload(decode=True)
            return body


        def extract_attachments(parsed_email):
            attachments = []
            for part in parsed_email.walk():
                if part.get_content_maintype() == 'multipart':
                    continue
                if part.get('Content-Disposition') is None:
                    continue

                filename = part.get_filename()
                if bool(filename):
                    attachments.append(File(filename, part.get_payload(decode=True)))
            return attachments

        import email
        raw_email = raw_email
        parsed_email = email.message_from_bytes(raw_email)
        subject = parsed_email['Subject']
        sender = parsed_email['From']
        recipient = parsed_email['To']
        date = parsed_email['Date']
        body = extract_body(parsed_email)
        attachments = extract_attachments(parsed_email)

        return cls(subject, sender, recipient, date, body, attachments)


    def save_attachments(self, folder):
        import os
        if not os.path.isdir(folder):
            os.mkdir(folder)

        for attachment in self.attachments:
            attachment.save(folder, attachment.name)

    def __str__(self):
        return f"Subject: {self.subject}\nFrom: {self.sender}\nTo: {self.recipient}\nDate: {self.date}\nBody: {self.body}\nAttachments: {self.attachments}"

class Schedule:
    def __init__(self, updateTime='', faculty='', field='', degree='', mode='', year='', semester='', group='', lessons=None):
        if lessons is None:
            lessons = []
        self.updateTime = updateTime
        self.faculty = faculty
        self.field = field
        self.degree = degree
        self.mode = mode
        self.year = year
        self.semester = semester
        self.group = group
        self.lessons: List[Lesson] = lessons

    @classmethod
    def get_timetables_from_file(cls, filename: str):
        if filename.endswith('.txt'):
            with open(filename, 'r') as f:
                return [cls.get_timetable_from_txt_data(f.read().splitlines())]
        elif filename.endswith('.xlsx'):
            with open(filename, 'rb') as f:
                return cls.get_timetables_from_xlsx_data_openpyxl(BytesIO(f.read()))
        else:
            return []

    @classmethod
    def get_timetables_from_xlsx_data_openpyxl(cls, data: BytesIO):
        from classes import Schedule, Lesson
        from openpyxl import load_workbook, cell

        def parse_sheet(sheet):
            schedules = []
            schedule_data = sheet.cell(row=1, column=3).value.split(',')

            fieldOfStudent = schedule_data[0].strip()
            degree = schedule_data[1].strip()
            year = schedule_data[2].split('Rok')[1].strip()
            semester = schedule_data[3].split('Semestr')[1].strip()
            updateTime = "00:00"

            time_row = sheet[3]
            max_row = sheet.max_row
            for i in range(4, max_row + 1):
                group_row = sheet[i]
                schedule = parse_row(group_row, time_row)

                if schedule.group == '':
                    continue

                schedule.updateTime = updateTime
                schedule.field = fieldOfStudent
                schedule.degree = degree
                schedule.year = year
                schedule.semester = semester
                schedule.mode = "ZO"
                schedule.faculty = "WZIM"

                schedules.append(schedule)

            return schedules

        def parse_row(row, time_row):
            schedule = Schedule.empty()
            max_col = len(row)
            groupName = str(row[1].value).strip()

            if groupName == "Grupy":
                return schedule

            schedule.group = groupName

            dayNum = 0
            dayText = str(row[0].value).strip()
            if 'Piątek' in dayText:
                dayNum = 5
            elif 'Sobota' in dayText:
                dayNum = 6
            elif 'Niedziela' in dayText:
                dayNum = 7

            lessons = []
            lesson = Lesson.empty()
            lesson.dayNumber = dayNum
            lessonRaw = ""

            for j in range(2, max_col):
                lessonRaw = str(row[j].value)
                if lessonRaw == 'None':
                    continue

                lesson.timeStart = time_row[j].value

                # Set j at the next Cell (not MergedCell)
                while(j+1 < max_col):
                    # If next cell is MergedCell, skip it
                    if type(row[j+1]) == cell.MergedCell:
                        j += 1
                    else:
                        break

                lesson.timeEnd = time_row[j].value

                lessonRaw = lessonRaw.split(',')
                lesson.name = lessonRaw[0].split('(')[0].strip()

                for text in lessonRaw:
                    if text == 'WARUNEK':
                        continue
                    elif '(' in text:
                        lesson.type = text.split('(')[1].split(')')[0].strip()
                    elif 's.' in text:
                        lesson.location = text.split('s.')[1].strip()
                    elif 'b.' in text:
                        lesson.building = text.split('b.')[1].split("]")[0].strip()
                    else:
                        isTeacher = True
                        for word in text.strip().split(' '):
                            if not word[0].isupper() or any(char.isdigit() for char in word):
                                isTeacher = False
                                break
                        if isTeacher:
                            lesson.teacher = text.strip()

                if not ']' in lessonRaw[-1]:
                    lesson.comment = lessonRaw[-1].strip().strip('.')

                lessons.append(lesson)
                lesson = Lesson.empty()

            schedule.lessons = lessons
            return schedule

        plans = []
        workbook = load_workbook(data, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            plans_for_sheet = parse_sheet(sheet)
            plans.extend(plans_for_sheet)

        return plans

    @classmethod
    def get_timetable_from_txt_data(cls, data: List[str]):
        # 10.10.2023 22:34
        # WZIM, 2023, Jesień , ST, Inf, inż, R4, S7, gr1, ISI-1 ;
        lessons = []
        # Get update time
        updateTime = data[0].strip() # 10.10.2023 22:34
        # Get name
        scheduleInfo = data[1].strip().split(',')
        faculty = scheduleInfo[0].strip() # WZIM
        year = scheduleInfo[1].strip() # 2023
        mode = scheduleInfo[3].strip() # ST
        fieldOfStudent = scheduleInfo[4].strip() # Inf
        degree = scheduleInfo[5].strip() # inż
        year = scheduleInfo[6].strip().removeprefix('R').strip() # R4
        semester = scheduleInfo[7].strip().removeprefix('S').strip() # S7
        group = scheduleInfo[8].strip().removeprefix('gr').strip() # gr1

        split_filter = '------------------------------------------------'
        blocks: List[List[str]] = []
        # Divide file into blocks of lessons
        block = []
        for line in data:
            if line == split_filter:
                blocks.append(block)
                block = []
            else:
                block.append(line)
        blocks.append(block)

        # Iterate over blocks
        # ZJ_01
        # Sztuczna Inteligencja [Lab]
        # d1, 10:30-12:00
        # 3/18
        # Aleksandra Konopka
        # U: 8 tygodni (ostatnie zajęcia 45 minut krócej)
        for block in blocks:
            if block == '' or "ZJ" not in block[0]:
                continue
            elif "end." in block:
                break

            # Parse text
            num = block[0].strip()
            name = block[1].split('[')[0].strip()
            type = block[1].split('[')[1].split(']')[0].strip()
            day = block[2].split(',')[0].strip().removeprefix('d').strip()
            timeStart = block[2].split(',')[1].split('-')[0].strip()
            timeEnd = block[2].split(',')[1].split('-')[1].strip()
            if block[3] == "zdalne":
                room = 0
                building = 0
            elif block[3].find('/') != -1:
                room = block[3].split('/')[0].strip()
                building = block[3].split('/')[1].strip()
            else:
                room = block[3].split(',')[0].strip()
                building = block[3].split(',')[1].strip()

            building = block[3].split('/')[0].strip()
            teacher = block[4].strip()
            comment = block[5].removeprefix('U:').strip()

            # Create lesson object
            lesson = Lesson(num=num, name=name, type=type, day=day, timeStart=timeStart, timeEnd=timeEnd, room=str(room), floor=building, building=building, teacher=teacher, comment=comment, created='', updated='')
            # Add lesson to lessons list
            lessons.append(lesson)
        return cls(updateTime, faculty, fieldOfStudent, degree, mode, year, semester, group, lessons)

    @classmethod
    def empty(cls):
        return cls()

    def __str__(self):
        return f"Update time: {self.updateTime}\nFaculty: {self.faculty}\nField: {self.field}\nDegree: {self.degree}\nMode: {self.mode}\nYear: {self.year}\nSemester: {self.semester}\nGroup: {self.group}\nLessons: {self.lessons}"

class Lesson:
    def __init__(self, num=0, created='', updated='', name='', type='', day=0, timeStart='', timeEnd='', room='', floor='', building='', teacher='', comment=''):
        self.num = num
        self.created = created
        self.updated = updated
        self.name = name
        self.type = type
        self.dayNumber = day
        self.timeStart = timeStart
        self.timeEnd = timeEnd
        self.location = room
        self.building = building
        self.teacher = teacher
        self.comment = comment

    @classmethod
    def empty(cls):
        return cls()

    def __str__(self):
        return f"{self.name}, {self.type}, {self.dayNumber}, {self.timeStart}, {self.timeEnd}, {self.location}, {self.teacher}"
